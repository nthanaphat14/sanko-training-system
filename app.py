import os
from io import BytesIO
from math import ceil
from datetime import datetime, date, timedelta
from functools import wraps

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    flash,
    session,
    g,
    abort,
    send_file,
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_user,
    logout_user,
    login_required,
)
from sqlalchemy import or_, func
from sqlalchemy.sql import nullslast
from sqlalchemy.exc import IntegrityError, DataError
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# -------------------------------------------------
# App Config
# -------------------------------------------------
app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "super-secret-key")

ALLOWED_ADMIN_EMAIL = "hr02@sankothai.net"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_EVENT_DIR = os.path.join(BASE_DIR, "static", "uploads", "events")
os.makedirs(UPLOAD_EVENT_DIR, exist_ok=True)

# Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message = "กรุณาเข้าสู่ระบบก่อน"

# Upload
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_COURSE_DIR = os.path.join(BASE_DIR, "static", "uploads", "courses")
os.makedirs(UPLOAD_COURSE_DIR, exist_ok=True)

ALLOWED_EXT = {"pdf", "png", "jpg", "jpeg", "xlsx"}

# Database
db_url = (os.environ.get("DATABASE_URL") or "").strip()
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = db_url or "sqlite:///employee.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = True
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(minutes=30)

print("DATABASE =", app.config["SQLALCHEMY_DATABASE_URI"])

db = SQLAlchemy(app)


# -------------------------------------------------
# Models
# -------------------------------------------------
class User(UserMixin, db.Model):
    __tablename__ = "users"

    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(180), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default="viewer")  # admin / viewer
    is_active = db.Column(db.Boolean, default=True)

    failed_attempts = db.Column(db.Integer, default=0)
    locked_until = db.Column(db.DateTime, nullable=True)

    last_login_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


class AuditLog(db.Model):
    __tablename__ = "audit_logs"

    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)

    user_email = db.Column(db.String(255), nullable=True, index=True)
    action = db.Column(db.String(80), nullable=False, index=True)
    detail = db.Column(db.Text, nullable=True)
    ip = db.Column(db.String(64), nullable=True)


class Employee(db.Model):
    __tablename__ = "employees"

    id = db.Column(db.Integer, primary_key=True)
    no = db.Column(db.Integer, nullable=True)
    em_id = db.Column(db.String(40), unique=True, nullable=False)
    id_card = db.Column(db.String(60), nullable=True)

    title_th = db.Column(db.String(50), nullable=True)
    first_name_th = db.Column(db.String(120), nullable=True)
    last_name_th = db.Column(db.String(120), nullable=True)

    title_en = db.Column(db.String(50), nullable=True)
    first_name_en = db.Column(db.String(120), nullable=True)
    last_name_en = db.Column(db.String(120), nullable=True)

    position = db.Column(db.String(150), nullable=True)
    section = db.Column(db.String(150), nullable=True)
    department = db.Column(db.String(150), nullable=True)

    start_work = db.Column(db.Date, nullable=True)
    resign = db.Column(db.Date, nullable=True)

    status = db.Column(db.String(50), nullable=True)
    degree = db.Column(db.String(80), nullable=True)
    major = db.Column(db.String(150), nullable=True)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def th_full(self):
        first = (self.first_name_th or "").strip()
        last = (self.last_name_th or "").strip()
        return f"{first} {last}".strip()

    def en_full(self):
        first = (self.first_name_en or "").strip()
        last = (self.last_name_en or "").strip()
        return f"{first} {last}".strip()


class TrainingRecord(db.Model):
    __tablename__ = "training_records"
    __table_args__ = (
        db.UniqueConstraint("emp_id", "start_date", name="uq_training_emp_start"),
    )

    id = db.Column(db.Integer, primary_key=True)

    seq = db.Column(db.Integer, nullable=True)
    year = db.Column(db.Integer, nullable=True)
    month = db.Column(db.Integer, nullable=True)

    emp_id = db.Column(db.String(50), nullable=False, index=True)

    prefix = db.Column(db.String(50), nullable=True)
    first_name = db.Column(db.String(200), nullable=True)
    last_name = db.Column(db.String(200), nullable=True)

    section = db.Column(db.String(150), nullable=True)
    position = db.Column(db.String(150), nullable=True)

    course_code = db.Column(db.String(100), nullable=True)
    course_name = db.Column(db.String(255), nullable=True)
    course_type = db.Column(db.String(100), nullable=True)

    start_date = db.Column(db.Date, nullable=True)
    end_date = db.Column(db.Date, nullable=True)
    hours = db.Column(db.Float, nullable=True)

    evaluate_method = db.Column(db.String(150), nullable=True)
    result = db.Column(db.String(50), nullable=True)
    score = db.Column(db.Float, nullable=True)
    evaluator = db.Column(db.String(150), nullable=True)

    expire_date = db.Column(db.Date, nullable=True)
    remark = db.Column(db.Text, nullable=True)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    event_id = db.Column(
        db.Integer,
        db.ForeignKey("training_events.id"),
        nullable=True,
        index=True
    )

    event = db.relationship(
        "TrainingEvent",
        backref=db.backref("records", lazy=True)
    )


class ImportBatch(db.Model):
    __tablename__ = "import_batches"

    id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    added = db.Column(db.Integer, default=0)
    updated = db.Column(db.Integer, default=0)
    duplicated = db.Column(db.Integer, default=0)
    skipped = db.Column(db.Integer, default=0)

    filename = db.Column(db.String(255), nullable=True)


class ImportItem(db.Model):
    __tablename__ = "import_items"

    id = db.Column(db.Integer, primary_key=True)
    batch_id = db.Column(db.Integer, db.ForeignKey("import_batches.id"), nullable=False, index=True)

    status = db.Column(db.String(20), nullable=False)
    reason = db.Column(db.String(255), nullable=True)

    row_no = db.Column(db.Integer, nullable=True)
    emp_id = db.Column(db.String(50), nullable=True)
    prefix = db.Column(db.String(50), nullable=True)
    first_name = db.Column(db.String(200), nullable=True)
    last_name = db.Column(db.String(200), nullable=True)
    section = db.Column(db.String(150), nullable=True)
    position = db.Column(db.String(150), nullable=True)

    course_code = db.Column(db.String(100), nullable=True)
    course_name = db.Column(db.String(255), nullable=True)
    course_type = db.Column(db.String(100), nullable=True)

    start_date = db.Column(db.Date, nullable=True)
    end_date = db.Column(db.Date, nullable=True)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class TrainingCourse(db.Model):
    __tablename__ = "training_courses"

    id = db.Column(db.Integer, primary_key=True)

    course_type = db.Column(db.String(20), nullable=False, index=True)  # OJT / INH / EXT
    course_code = db.Column(db.String(30), nullable=False, unique=True, index=True)

    course_name = db.Column(db.String(255), nullable=False)
    description = db.Column(db.Text, nullable=True)

    owner = db.Column(db.String(120), nullable=True)
    vendor = db.Column(db.String(200), nullable=True)
    location = db.Column(db.String(200), nullable=True)
    training_hours = db.Column(db.Float, nullable=True)
    training_date = db.Column(db.Date, nullable=True)

    course_year = db.Column(db.Integer, nullable=True, index=True)
    course_month = db.Column(db.Integer, nullable=True, index=True)

    status = db.Column(db.String(30), default="Draft")

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    def total_before_vat(self):
        return float(sum(c.amount_before_vat or 0 for c in self.cost_items))

    def total_vat(self):
        return float(sum(c.amount_vat or 0 for c in self.cost_items))

    def total_amount(self):
        return float(sum(c.amount_total or 0 for c in self.cost_items))


class CourseFile(db.Model):
    __tablename__ = "course_files"

    id = db.Column(db.Integer, primary_key=True)
    course_id = db.Column(db.Integer, db.ForeignKey("training_courses.id"), nullable=False, index=True)

    file_type = db.Column(db.String(50), nullable=True)
    original_name = db.Column(db.String(255), nullable=True)
    stored_name = db.Column(db.String(255), nullable=False)
    note = db.Column(db.String(255), nullable=True)

    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)

    course = db.relationship("TrainingCourse", backref=db.backref("files", lazy=True))


class CourseCostItem(db.Model):
    __tablename__ = "course_cost_items"

    id = db.Column(db.Integer, primary_key=True)
    course_id = db.Column(db.Integer, db.ForeignKey("training_courses.id"), nullable=False, index=True)

    cost_type = db.Column(db.String(80), nullable=False)
    amount_before_vat = db.Column(db.Float, nullable=True)
    vat_rate = db.Column(db.Float, default=7.0)
    amount_vat = db.Column(db.Float, nullable=True)
    amount_total = db.Column(db.Float, nullable=True)
    remark = db.Column(db.String(255), nullable=True)

    attach_file_id = db.Column(db.Integer, db.ForeignKey("course_files.id"), nullable=True)
    attach_file = db.relationship("CourseFile", foreign_keys=[attach_file_id])

    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    course = db.relationship("TrainingCourse", backref=db.backref("cost_items", lazy=True))

class EventCostItem(db.Model):
    __tablename__ = "event_cost_items"

    id = db.Column(db.Integer, primary_key=True)

    event_id = db.Column(
        db.Integer,
        db.ForeignKey("training_events.id"),
        nullable=False,
        index=True
    )

    cost_type = db.Column(db.String(80), nullable=False)
    amount_before_vat = db.Column(db.Float, nullable=True)
    vat_rate = db.Column(db.Float, default=7.0)
    amount_vat = db.Column(db.Float, nullable=True)
    amount_total = db.Column(db.Float, nullable=True)
    remark = db.Column(db.String(255), nullable=True)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    event = db.relationship(
        "TrainingEvent",
        backref=db.backref("cost_items", lazy=True, cascade="all, delete-orphan")
    )

class TrainingEvent(db.Model):
    __tablename__ = "training_events"

    id = db.Column(db.Integer, primary_key=True)

    course_id = db.Column(
        db.Integer,
        db.ForeignKey("training_courses.id"),
        nullable=False,
        index=True
    )

    event_type = db.Column(db.String(20), nullable=False, index=True)
    event_code = db.Column(db.String(30), nullable=False, unique=True, index=True)

    title = db.Column(db.String(255), nullable=False)
    location = db.Column(db.String(255), nullable=True)

    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=True)

    trainer = db.Column(db.String(255), nullable=True)

    description = db.Column(db.Text, nullable=True)
    status = db.Column(db.String(30), nullable=False, default="PLANNED")

    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    course = db.relationship(
        "TrainingCourse",
        backref=db.backref("events", lazy=True)
    )

class TrainingEventParticipant(db.Model):
    __tablename__ = "training_event_participants"

    id = db.Column(db.Integer, primary_key=True)

    event_id = db.Column(
        db.Integer,
        db.ForeignKey("training_events.id"),
        nullable=False,
        index=True
    )

    emp_id = db.Column(db.String(40), nullable=False, index=True)

    # ✅ เพิ่ม 4 ช่องนี้
    result = db.Column(db.String(20), nullable=True)
    score = db.Column(db.Float, nullable=True)
    training_hours = db.Column(db.Float, nullable=True)
    remark = db.Column(db.String(255), nullable=True)

    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    event = db.relationship(
        "TrainingEvent",
        backref=db.backref("participants", lazy=True, cascade="all, delete-orphan")
    )

class EventFile(db.Model):
    __tablename__ = "event_files"

    id = db.Column(db.Integer, primary_key=True)

    event_id = db.Column(
        db.Integer,
        db.ForeignKey("training_events.id"),
        nullable=False,
        index=True
    )

    file_type = db.Column(db.String(50), nullable=True)   # quotation / invoice / receipt / certificate / attendance / other
    original_name = db.Column(db.String(255), nullable=True)
    stored_name = db.Column(db.String(255), nullable=False)
    note = db.Column(db.String(255), nullable=True)

    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)

    event = db.relationship(
        "TrainingEvent",
        backref=db.backref("files", lazy=True, cascade="all, delete-orphan")
    )
# -------------------------------------------------
# Helper Functions
# -------------------------------------------------
def allowed_file(filename: str) -> bool:
    if not filename or "." not in filename:
        return False
    ext = filename.rsplit(".", 1)[1].lower()
    return ext in ALLOWED_EXT

import re

def normalize_owner_code(owner: str | None) -> str:
    owner = (owner or "").strip().upper()
    owner = re.sub(r"[^A-Z0-9]", "", owner)
    return owner or "GEN"

def gen_course_code(course_type: str, owner: str | None = None, dt: datetime | None = None) -> str:
    dt = dt or datetime.utcnow()
    year = dt.year
    prefix = (course_type or "").strip().upper()
    owner_code = normalize_owner_code(owner)

    like = f"{prefix}-{year}-{owner_code}-%"

    last_code = db.session.query(func.max(TrainingCourse.course_code)).filter(
        TrainingCourse.course_code.ilike(like)
    ).scalar()

    if last_code:
        try:
            last_run = int(last_code.split("-")[-1])
        except Exception:
            last_run = 0
    else:
        last_run = 0

    new_run = last_run + 1
    return f"{prefix}-{year}-{owner_code}-{new_run:04d}"



def get_current_user():
    uid = session.get("uid")
    if not uid:
        return None
    return db.session.get(User, uid)

def audit(action, detail=""):
    """
    บันทึกการใช้งานลง audit_logs
    action: ชื่อเหตุการณ์ เช่น LOGIN_SUCCESS, EMPLOYEE_ADD
    detail: รายละเอียดเพิ่มเติม
    """
    try:
        # ดึง email จาก session ก่อน
        user_email = session.get("user_email")

        # ถ้าไม่มีใน session ให้ลองดึงจาก user ปัจจุบัน
        if not user_email:
            u = get_current_user()
            if u:
                user_email = u.email

        # IP address
        ip = request.headers.get("X-Forwarded-For", request.remote_addr)
        if ip and "," in ip:
            ip = ip.split(",")[0].strip()

        log = AuditLog(
            user_email=user_email,
            action=action,
            detail=detail,
            ip=ip,
        )
        db.session.add(log)
        db.session.commit()

    except Exception as e:
        # กันระบบหลักพังเพราะ log
        db.session.rollback()
        print("AUDIT LOG ERROR:", e)

def build_training_query(args):
    q = (args.get("q") or "").strip()
    year = (args.get("year") or "").strip()
    month = (args.get("month") or "").strip()

    query = TrainingRecord.query

    if q:
        like = f"%{q}%"
        name_field = None

        for cand in ["full_name", "employee_name", "name", "emp_name", "first_name", "last_name"]:
            if hasattr(TrainingRecord, cand):
                name_field = getattr(TrainingRecord, cand)
                break

        conds = []

        if hasattr(TrainingRecord, "emp_id"):
            conds.append(TrainingRecord.emp_id.ilike(like))
        if hasattr(TrainingRecord, "employee_code"):
            conds.append(TrainingRecord.employee_code.ilike(like))
        if hasattr(TrainingRecord, "first_name"):
            conds.append(TrainingRecord.first_name.ilike(like))
        if hasattr(TrainingRecord, "last_name"):
            conds.append(TrainingRecord.last_name.ilike(like))
        if name_field is not None:
            conds.append(name_field.ilike(like))
        if hasattr(TrainingRecord, "course_code"):
            conds.append(TrainingRecord.course_code.ilike(like))
        if hasattr(TrainingRecord, "course_name"):
            conds.append(TrainingRecord.course_name.ilike(like))

        if conds:
            query = query.filter(or_(*conds))

    if year.isdigit() and hasattr(TrainingRecord, "year"):
        query = query.filter(TrainingRecord.year == int(year))
    if month.isdigit() and hasattr(TrainingRecord, "month"):
        query = query.filter(TrainingRecord.month == int(month))

    return query, q, year, month


@app.before_request
def block_non_owner():
    public_endpoints = {
        "login",
        "logout",
        "static",
        "robots",
        "favicon",
    }

    if request.endpoint is None:
        return

    if request.endpoint in public_endpoints:
        return

    try:
        if not current_user.is_authenticated:
            return
    except Exception:
        return

    user_email = getattr(current_user, "email", None)
    if not user_email:
        abort(403)

    if user_email.strip().lower() != ALLOWED_ADMIN_EMAIL.lower():
        abort(403)


@app.route("/favicon.ico")
def favicon():
    return "", 204
    
@app.context_processor
def inject_helpers():
    return {
        "current_user": get_current_user(),
        "mask_id_card": mask_id_card,
        "format_id_card": format_id_card,
    }

def format_id_card(id_card: str):
    if not id_card or len(id_card) != 13:
        return ""
    return f"{id_card[0]}-{id_card[1:5]}-{id_card[5:10]}-{id_card[10:12]}-{id_card[12]}"

def mask_id_card(id_card):
    if not id_card:
        return ""
    return id_card[:6] + "XXXXX" + id_card[-3:]

app.jinja_env.globals.update(mask_id_card=mask_id_card)

def audit(action, detail=None, user_email=None):
    try:
        ip = (request.headers.get("X-Forwarded-For") or request.remote_addr or "")
        ip = ip.split(",")[0].strip() if ip else None

        if user_email is None:
            u = get_current_user()  # ถ้ามี session อยู่จะได้ email
            user_email = getattr(u, "email", None) if u else None

        db.session.add(AuditLog(
            action=str(action),
            detail=detail,
            user_email=user_email,
            ip=ip,
        ))
        db.session.commit()
    except Exception:
        db.session.rollback()
        # ห้ามให้ audit ทำให้ระบบล่ม
        pass

@app.before_request
def require_login_globally():
    # บางครั้ง endpoint เป็น None (เช่น 404) กันพังไว้
    if request.endpoint is None:
        return

    allow = {"login", "login_post", "logout", "static"}

    if request.endpoint in allow:
        return

    u = get_current_user()
    if not u or not u.is_active:
        return redirect(url_for("login"))
        
def role_required(*roles):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            u = get_current_user()
            if not u or not u.is_active:
                return redirect(url_for("login"))

            if roles:
                if getattr(u, "role", None) not in roles:
                    abort(403)

            return fn(*args, **kwargs)
        return wrapper
    return deco

def seed_users_if_missing():
    defaults = [
        ("hr02@sankothai.net", "Sanko1996", "admin"),
        ("hr@sankothai.net", "Sanko1996", "viewer"),
        ("hr01@sankothai.net", "Sanko1996", "viewer"),
    ]

    for email, pw, role in defaults:
        exists = User.query.filter_by(email=email).first()
        if not exists:
            db.session.add(User(
                email=email,
                password_hash=generate_password_hash(pw),
                role=role,
                is_active=True
            ))

    db.session.commit()

def login_required(fn):
    # ใช้ role_required แบบไม่กำหนด role (แค่ต้อง login)
    return role_required()(fn)
    
def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()
    
def build_employee_query(q="", status="Active", dept="", section="", sort="no", direction="asc"):
    query = Employee.query

    # ---- status filter ----
    if status and status != "All":
        query = query.filter(Employee.status == status)

    # ---- dept/section filter ----
    if dept:
        query = query.filter(Employee.department == dept)
    if section:
        query = query.filter(Employee.section == section)

    # ---- search filter ----
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Employee.em_id.ilike(like),
                Employee.id_card.ilike(like),
                Employee.first_name_th.ilike(like),
                Employee.last_name_th.ilike(like),
                Employee.first_name_en.ilike(like),
                Employee.last_name_en.ilike(like),
                Employee.position.ilike(like),
                Employee.section.ilike(like),
                Employee.department.ilike(like),
            )
        )

    # ---- sorting ----
    sort_map = {
        "no": Employee.no,
        "em_id": Employee.em_id,
        "department": Employee.department,
        "section": Employee.section,
    }
    sort_col = sort_map.get(sort or "no", Employee.no)

    if (direction or "asc").lower() == "desc":
        query = query.order_by(nullslast(sort_col.desc()))
    else:
        query = query.order_by(nullslast(sort_col.asc()))

    # tie-breaker ให้เสถียร
    query = query.order_by(Employee.em_id.asc())

    return query

def safe_int(x):
    try:
        if x is None: return None
        s = str(x).strip()
        if s == "": return None
        return int(float(s))
    except:
        return None
        
def safe_month(v):
    if v is None:
        return None

    # ถ้าเป็นเลขอยู่แล้ว
    try:
        iv = int(v)
        if 1 <= iv <= 12:
            return iv
    except:
        pass

    s = str(v).strip().lower()
    m = {
        "jan": 1, "january": 1,
        "feb": 2, "february": 2,
        "mar": 3, "march": 3,
        "apr": 4, "april": 4,
        "may": 5,
        "jun": 6, "june": 6,
        "jul": 7, "july": 7,
        "aug": 8, "august": 8,
        "sep": 9, "sept": 9, "september": 9,
        "oct": 10, "october": 10,
        "nov": 11, "november": 11,
        "dec": 12, "december": 12,
    }
    return m.get(s)       

def safe_date(v):
    """
    รับได้ทั้ง:
    - datetime/date
    - ตัวเลข excel date (บางครั้ง openpyxl จะให้เป็น datetime อยู่แล้ว)
    - string หลาย format เช่น 2026-02-25, 25/02/2026, 25-02-2026
    """
    if v is None:
        return None

    # ถ้าเป็น date/datetime อยู่แล้ว
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = str(v).strip()
    if not s or s.lower() in ("none", "nan"):
        return None

    # ลอง format ที่พบบ่อย
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except:
            pass

    return None
    
MONTH_MAP = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}

def month_to_int(v):
    s = safe_str(v).strip().lower()
    if s.isdigit():
        return int(s)
    parse(s[:3]) or MONTH_MAP.get(s)

def safe_float(x):
    try:
        if x is None: return None
        s = str(x).strip()
        if s == "": return None
        return float(s)
    except:
        return None

def safe_date(v):
    # รองรับ date/datetime จาก Excel + string หลายรูปแบบ
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except:
            pass
    return None

def gen_event_code(event_type: str, event_date: date | None = None) -> str:
    event_date = event_date or datetime.utcnow().date()
    prefix = (event_type or "").strip().upper()
    ymd = event_date.strftime("%Y%m%d")

    like = f"{prefix}-{ymd}-%"

    last_code = db.session.query(func.max(TrainingEvent.event_code)).filter(
        TrainingEvent.event_code.ilike(like)
    ).scalar()

    if last_code:
        try:
            last_run = int(last_code.split("-")[-1])
        except Exception:
            last_run = 0
    else:
        last_run = 0

    return f"{prefix}-{ymd}-{last_run + 1:04d}"
    
# -------------------------------------------------
# Routes
# -------------------------------------------------

@app.get("/login")
def login():
    u = get_current_user()
    if u and u.is_active:
        return redirect(url_for("employees_list"))
    return render_template("login.html")


@app.post("/login")
def login_post():
    email = (request.form.get("email") or "").strip().lower()
    password = request.form.get("password") or ""

    u = User.query.filter_by(email=email).first()

    if not u or not u.is_active:
        audit("LOGIN_FAIL", f"email={email}")
        flash("User หรือ Password ไม่ถูกต้อง", "error")
        return redirect(url_for("login"))

    # ถ้าโดน lock
    if u.locked_until and u.locked_until > datetime.utcnow():
        audit("LOGIN_LOCKED", f"email={email}")
        flash("บัญชีถูกล็อกชั่วคราว", "error")
        return redirect(url_for("login"))

    # password ไม่ถูก
    if not check_password_hash(u.password_hash, password):
        u.failed_attempts = (u.failed_attempts or 0) + 1

        if u.failed_attempts >= 8:
            u.locked_until = datetime.utcnow() + timedelta(minutes=10)
            u.failed_attempts = 0

        db.session.commit()

        audit("LOGIN_FAIL", f"email={email}")
        flash("User หรือ Password ไม่ถูกต้อง", "error")
        return redirect(url_for("login"))

    # LOGIN SUCCESS
    u.failed_attempts = 0
    u.locked_until = None
    u.last_login_at = datetime.utcnow()

    db.session.commit()

    session.clear()
    session["uid"] = u.id
    session["user_email"] = u.email
    session["user_role"] = u.role
    session.permanent = True

    audit("LOGIN_SUCCESS", f"email={email}")

    flash("เข้าสู่ระบบสำเร็จ", "success")
    return redirect(url_for("employees_list"))


@app.get("/logout")
def logout():
    audit("LOGOUT", "")
    session.clear()
    return redirect(url_for("login"))


# -------------------------------------------------
# Change Password
# -------------------------------------------------

@app.route("/change-password", methods=["GET", "POST"])
def change_password():
    u = get_current_user()

    if not u:
        return redirect(url_for("login"))

    if request.method == "GET":
        return render_template("change_password.html", user=u)

    old_pw = request.form.get("old_password") or ""
    new_pw = request.form.get("new_password") or ""
    new_pw2 = request.form.get("new_password2") or ""

    if not check_password_hash(u.password_hash, old_pw):
        flash("รหัสเดิมไม่ถูกต้อง", "error")
        return redirect(url_for("change_password"))

    if new_pw != new_pw2:
        flash("ยืนยันรหัสใหม่ไม่ตรงกัน", "error")
        return redirect(url_for("change_password"))

    if len(new_pw) < 10:
        flash("รหัสใหม่ต้องยาวอย่างน้อย 10 ตัวอักษร", "error")
        return redirect(url_for("change_password"))

    u.password_hash = generate_password_hash(new_pw)

    db.session.commit()

    audit("CHANGE_PASSWORD", f"user={u.email}")

    flash("เปลี่ยนรหัสผ่านสำเร็จ", "success")
    return redirect(url_for("employees_list"))


# -------------------------------------------------
# Admin Users
# -------------------------------------------------

@app.get("/admin/users")
def admin_users():
    u = get_current_user()

    if not u or u.role != "admin":
        abort(403)

    users = User.query.order_by(User.email.asc()).all()
    return render_template("admin_users.html", users=users)


@app.post("/admin/users/<int:user_id>/reset-password")
def admin_reset_password(user_id):

    u = get_current_user()
    if not u or u.role != "admin":
        abort(403)

    target = User.query.get_or_404(user_id)
    new_pw = request.form.get("new_password") or ""

    if len(new_pw) < 10:
        flash("รหัสใหม่ต้องยาวอย่างน้อย 10 ตัวอักษร", "error")
        return redirect(url_for("admin_users"))

    target.password_hash = generate_password_hash(new_pw)

    db.session.commit()

    audit("RESET_PASSWORD", f"admin={u.email} reset={target.email}")

    flash(f"รีเซ็ตรหัสผ่านให้ {target.email} สำเร็จ", "success")

    return redirect(url_for("admin_users"))


# -------------------------------------------------
# Require login globally
# -------------------------------------------------

@app.before_request
def require_login_globally():

    open_paths = {"/login", "/healthz"}

    if request.path.startswith("/static/"):
        return

    if request.path in open_paths:
        return

    if request.path == "/login" and request.method == "POST":
        return

    u = get_current_user()

    if not u or not u.is_active:
        return redirect(url_for("login"))


# -------------------------------------------------
# Root
# -------------------------------------------------

@app.get("/")
def root():
    return redirect(url_for("employees_list"))


@app.get("/healthz")
def healthz():
    return {"status": "ok"}, 200

@app.get("/employees")
@login_required
def employees_list():
    q = (request.args.get("q") or "").strip()

    # ✅ รับค่า filter/sort จาก URL
    status = (request.args.get("status") or "Active").strip()
    dept = (request.args.get("dept") or "").strip()
    section = (request.args.get("section") or "").strip()
    sort = (request.args.get("sort") or "no").strip()
    direction = (request.args.get("direction") or "asc").strip().lower()

    query = Employee.query

    # ✅ Filter: Status
    if status in ["Active", "Resign"]:
        query = query.filter(Employee.status == status)
    # ถ้าเป็น All หรือค่าอื่น → ไม่กรอง

    # ✅ Filter: Department / Section
    if dept:
        query = query.filter(Employee.department == dept)
    if section:
        query = query.filter(Employee.section == section)

    # ✅ Search
    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Employee.em_id.ilike(like),
                Employee.id_card.ilike(like),
                Employee.first_name_th.ilike(like),
                Employee.last_name_th.ilike(like),
                Employee.first_name_en.ilike(like),
                Employee.last_name_en.ilike(like),
                Employee.position.ilike(like),
                Employee.section.ilike(like),
                Employee.department.ilike(like),
                Employee.status.ilike(like),
            )
        )

    # ✅ Dropdown options (ต้องเป็น list ของ string)
    dept_options = [
        d[0] for d in db.session.query(Employee.department)
        .filter(Employee.department.isnot(None))
        .filter(Employee.department != "")
        .distinct()
        .order_by(Employee.department.asc())
        .all()
    ]

    section_options = [
        s[0] for s in db.session.query(Employee.section)
        .filter(Employee.section.isnot(None))
        .filter(Employee.section != "")
        .distinct()
        .order_by(Employee.section.asc())
        .all()
    ]

    # ✅ Sort
    sort_map = {
        "no": Employee.no,
        "em_id": Employee.em_id,
        "department": Employee.department,
        "section": Employee.section,
    }
    sort_col = sort_map.get(sort, Employee.no)

    if direction == "desc":
        query = query.order_by(nullslast(sort_col.desc()))
    else:
        query = query.order_by(nullslast(sort_col.asc()))

    # เสริมเรียง em_id ต่อท้ายกันข้อมูลกระโดด
    if sort != "em_id":
        query = query.order_by(
            nullslast(sort_col.desc() if direction == "desc" else sort_col.asc()),
            Employee.em_id.asc()
        )

    employees = query.all()

    # ✅ Summary counts (นับจากฐานจริง ไม่ใช่จาก filter)
    total_active = Employee.query.filter(Employee.status == "Active").count()
    total_resign = Employee.query.filter(Employee.status == "Resign").count()
    total_all = Employee.query.count()

    return render_template(
        "employees.html",
        employees=employees,
        total=len(employees),
        q=q,

        status=status,
        dept=dept,
        section=section,
        sort=sort,
        direction=direction,

        dept_options=dept_options,
        section_options=section_options,

        total_active=total_active,
        total_resign=total_resign,
        total_all=total_all,
    )
    
@app.route("/employees/new", methods=["GET", "POST"])
def employee_new():
    if request.method == "POST":
        em_id = request.form.get("em_id", "").strip()
        first_name_th = request.form.get("first_name_th", "").strip()

        if not em_id:
            flash("กรุณากรอกรหัสพนักงาน", "error")
            return redirect(url_for("employee_new"))

        if Employee.query.filter_by(em_id=em_id).first():
            flash("em_id นี้มีอยู่แล้ว", "error")
            return redirect(url_for("employee_new"))

        emp = Employee(em_id=em_id, first_name_th=first_name_th)
        db.session.add(emp)
        db.session.commit()
        audit("EMPLOYEE_ADD", f"em_id={emp.em_id}")

        flash("เพิ่มพนักงานเรียบร้อย", "success")
        return redirect(url_for("employees_list"))

    return render_template("employee_form.html", employee=None)

from sqlalchemy.exc import IntegrityError, DataError

def normalize_status(x):
    if not x:
        return None
    s = str(x).strip().upper()
    if s in ("W", "WORKING", "ACTIVE", "ทำงาน", "ยังอยู่"):
        return "W"
    if s in ("RS", "RESIGN", "RESIGNED", "ลาออก", "ออก"):
        return "RS"
    return s

@app.route("/employees/<string:em_id>/edit", methods=["GET", "POST"])
def employee_edit(em_id):
    emp = Employee.query.filter_by(em_id=em_id).first_or_404()

    if request.method == "POST":
        try:
            emp.id_card = safe_str(request.form.get("id_card"))
            emp.first_name_th = safe_str(request.form.get("first_name_th"))
            emp.last_name_th  = safe_str(request.form.get("last_name_th"))
            emp.first_name_en = safe_str(request.form.get("first_name_en"))
            emp.last_name_en  = safe_str(request.form.get("last_name_en"))
            emp.status = normalize_status(request.form.get("status"))

            db.session.commit()
            flash("แก้ไขข้อมูลเรียบร้อย", "success")
            return redirect(url_for("employees_list"))

        except (IntegrityError, DataError) as e:
            db.session.rollback()
            audit("EMPLOYEE_EDIT", f"em_id={emp.em_id}")
            
            flash("บันทึกไม่สำเร็จ: ข้อมูลไม่ถูกต้องหรือซ้ำในระบบ", "error")
        except Exception as e:
            db.session.rollback()
            flash(f"เกิดข้อผิดพลาด: {e}", "error")
    return render_template("employee_form.html", employee=emp)

@app.route("/employees/<string:em_id>/delete", methods=["POST"])
def employee_delete(em_id):
    emp = Employee.query.filter_by(em_id=em_id).first_or_404()
    db.session.delete(emp)
    db.session.commit()
    em_id = emp.em_id
    audit("EMPLOYEE_EDIT", f"em_id={emp.em_id}")
    
    flash("ลบข้อมูลเรียบร้อย", "success")
    return redirect(url_for("employees_list"))

@app.post("/employees/bulk-delete")
@role_required("admin")
def employees_bulk_delete():
    ids = request.form.getlist("ids")

    if not ids:
        flash("ยังไม่ได้เลือกพนักงาน", "error")
        return redirect(url_for("employees_list"))

    Employee.query.filter(Employee.em_id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()

    audit("EMPLOYEE_BULK_DELETE", f"count={len(ids)}")
    flash(f"ลบแล้ว {len(ids)} รายการ", "success")
    return redirect(url_for("employees_list"))

@app.route("/employees/import", methods=["GET", "POST"])
@role_required("admin")
def employees_import():
    if request.method == "GET":
        return render_template("employees_import.html")

    f = request.files.get("file")
    if not f:
        flash("กรุณาเลือกไฟล์ .xlsx", "error")
        return redirect(url_for("employees_import"))

    import re

    try:
        wb = load_workbook(f, data_only=True)

        added = 0
        updated = 0
        skipped = 0

        for ws in wb.worksheets:
            sheet_title = (ws.title or "").strip().lower()
            default_status = "Resign" if ("ลาออก" in sheet_title or "resign" in sheet_title) else "Active"

            # -------- หาแถวหัวตารางที่มี Em. ID --------
            header_row = None
            for r in range(1, 30):
                values = [(str(c.value).strip() if c.value is not None else "") for c in ws[r]]
                if any(("em. id" in v.lower() or "em id" in v.lower()) for v in values):
                    header_row = r
                    break

            if header_row is None:
                # ชีตนี้ไม่ใช่ employee data -> ข้าม
                continue

            headers = [(str(c.value).strip() if c.value is not None else "") for c in ws[header_row]]

            def col(*names):
                """คืน index ของคอลัมน์จากรายชื่อที่เป็นไปได้หลายแบบ"""
                header_l = [(h or "").strip().lower() for h in headers]
                for name in names:
                    key = (name or "").strip().lower()
                    if key in header_l:
                        return header_l.index(key)
                return None

            # -------- map คอลัมน์ (รองรับหลายชื่อ) --------
            c_no       = col("no.", "no")
            c_em       = col("em. id", "em id", "employee id", "emp id")
            c_prefix   = col("prefix")
            c_first_th = col("name-th", "name th", "first-th", "first th")
            c_last_th  = col("last-th", "last th", "surname-th", "surname th")
            c_name_en  = col("name-en", "name en")

            c_position = col("position")
            c_section  = col("section")
            c_dept     = col("department", "dept")

            c_start    = col("hire date", "start work", "start date")
            c_resign   = col("resign", "resign date", "end date")

            c_idcard   = col("id card", "idcard", "citizen id", "citizenid")
            c_degree   = col("education", "degree")
            c_major    = col("major")
            c_school   = col("school name", "institute", "university")

            if c_em is None:
                continue

            # -------- วนอ่านข้อมูลหลังหัวตาราง --------
            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                # กันแถวสั้น/ว่าง
                if not row or c_em >= len(row):
                    skipped += 1
                    continue

                em_id = safe_str(row[c_em]).strip() if row[c_em] is not None else ""
                if not em_id:
                    skipped += 1
                    continue

                # รองรับรหัสขึ้นต้นด้วย M หรือ D เท่านั้น (ปรับจำนวนหลักได้)
                if not re.match(r"^[MD]\d{5,10}$", em_id, re.IGNORECASE):
                    skipped += 1
                    continue

                emp = Employee.query.filter_by(em_id=em_id).first()
                if not emp:
                    emp = Employee(em_id=em_id)
                    db.session.add(emp)
                    added += 1
                else:
                    updated += 1

                # ---- set ค่า (ทับตามไฟล์ Admin) ----
                emp.no = safe_int(row[c_no]) if c_no is not None and c_no < len(row) else None

                emp.title_th      = safe_str(row[c_prefix])   if c_prefix is not None and c_prefix < len(row) else ""
                emp.first_name_th = safe_str(row[c_first_th]) if c_first_th is not None and c_first_th < len(row) else ""
                emp.last_name_th  = safe_str(row[c_last_th])  if c_last_th is not None and c_last_th < len(row) else ""

                # อังกฤษในไฟล์เป็น full name ก้อนเดียว -> เก็บไว้ใน first_name_en (แบบที่คุณใช้อยู่)
                emp.first_name_en = safe_str(row[c_name_en]) if c_name_en is not None and c_name_en < len(row) else ""

                emp.position   = safe_str(row[c_position]) if c_position is not None and c_position < len(row) else ""
                emp.section    = safe_str(row[c_section])  if c_section is not None and c_section < len(row) else ""
                emp.department = safe_str(row[c_dept])     if c_dept is not None and c_dept < len(row) else ""

                emp.start_work = safe_date(row[c_start])  if c_start is not None and c_start < len(row) else None
                emp.resign     = safe_date(row[c_resign]) if c_resign is not None and c_resign < len(row) else None

                emp.id_card = safe_str(row[c_idcard]) if c_idcard is not None and c_idcard < len(row) else ""

                emp.degree = safe_str(row[c_degree]) if c_degree is not None and c_degree < len(row) else ""
                emp.major  = safe_str(row[c_major])  if c_major is not None and c_major < len(row) else ""

                # ถ้าใน Model คุณมีฟิลด์ school_name ให้เปิดใช้บรรทัดนี้
                if hasattr(emp, "school_name"):
                    emp.school_name = safe_str(row[c_school]) if c_school is not None and c_school < len(row) else ""

                # status: ใช้จากชื่อชีตเป็นค่าเริ่มต้น แต่ถ้ามี resign date ให้เป็น Resign
                emp.status = "Resign" if emp.resign else default_status

        db.session.commit()
        flash(f"Import Employees สำเร็จ: เพิ่มใหม่ {added} | อัปเดต {updated} | ข้าม {skipped}", "success")
        return redirect(url_for("employees_list"))

    except Exception as e:
        db.session.rollback()
        flash(f"Import Employees ล้มเหลว: {e}", "error")
        return redirect(url_for("employees_import"))
    
@app.route("/trainings/import", methods=["GET", "POST"])
@role_required("admin")
def trainings_import():
    if request.method == "GET":
        return render_template("training_import.html")

    f = request.files.get("file")
    if not f or f.filename == "":
        flash("กรุณาเลือกไฟล์ Excel", "error")
        return redirect(url_for("trainings_import"))

    # ---- สร้าง Batch รอบนี้ ----
    batch = ImportBatch(filename=f.filename)
    db.session.add(batch)
    db.session.commit()  # ต้อง commit ก่อนเพื่อได้ batch.id

    def log_item(status, reason=None, row_no=None, **kw):
        it = ImportItem(
            batch_id=batch.id,
            status=status,
            reason=reason,
            row_no=row_no,
            emp_id=kw.get("emp_id"),
            prefix=kw.get("prefix"),
            first_name=kw.get("first_name"),
            last_name=kw.get("last_name"),
            section=kw.get("section"),
            position=kw.get("position"),
            course_code=kw.get("course_code"),
            course_name=kw.get("course_name"),
            course_type=kw.get("course_type"),
            start_date=kw.get("start_date"),
            end_date=kw.get("end_date"),
        )
        db.session.add(it)

    try:
        wb = load_workbook(f, data_only=True)
        ws = wb["Record Training"] if "Record Training" in wb.sheetnames else wb.active

        # ======= NORMALIZE HEADER =======
        def norm(x):
            s = safe_str(x).strip().lower()
            for ch in ["\u00a0", ".", "-", "_", "/", "(", ")", "[", "]", ":"]:
                s = s.replace(ch, " ")
            return " ".join(s.split())

        # ✅ เพิ่ม alias ให้ตรงไฟล์คุณ (ชื่อไทย / name-th ฯลฯ)
        ALIASES = {
            "seq": ["ลำดับ", "no", "seq", "#"],
            "year": ["year", "year.", "ปี"],
            "month": ["month", "mon", "เดือน"],

            "emp_id": ["em id", "emp id", "empid", "รหัสพนักงาน", "รหัส", "employee id"],
            "prefix": ["คำนำหน้า", "prefix"],
            "first_name": ["ชื่อ", "ชื่อไทย", "name th", "thai name", "first name", "firstname"],
            "last_name": ["นามสกุล", "last name", "lastname"],
            "section": ["section", "แผนก", "ฝ่าย", "หน่วยงาน", "department"],
            "position": ["position", "ตำแหน่ง"],

            "course_code": ["course code", "coursecode", "รหัสหลักสูตร"],
            "course_name": ["course name", "coursename", "ชื่อหลักสูตร"],
            "course_type": ["course type", "type", "category", "ประเภท"],

            "start_date": ["startdate", "start date", "วันที่เริ่ม", "เริ่ม"],
            "end_date": ["enddate", "end date", "วันที่จบ", "จบ"],
            "hours": ["ชั่วโมง", "hours", "hour"],

            "evaluate_method": ["วิธีประเมิน", "evaluate method", "ประเมิน"],
            "result": ["ผล", "result"],
            "score": ["คะแนน", "score"],
            "evaluator": ["ผู้ประเมิน", "evaluator"],
            "expire_date": ["วันหมดอายุ", "expire date", "expiry"],
            "remark": ["หมายเหตุ", "remark", "note"],
        }

        def find_header_row(scan_rows=10):
            best_row, best_score = 1, -1
            max_r = min(scan_rows, ws.max_row or 1)
            max_c = ws.max_column or 1
            must_keys = ["emp_id", "course_code", "course_name", "start_date"]

            for r in range(1, max_r + 1):
                vals = [norm(ws.cell(r, c).value) for c in range(1, max_c + 1)]
                score = 0
                for k in must_keys:
                    if any(norm(a) in vals for a in ALIASES[k]):
                        score += 1
                if score > best_score:
                    best_score, best_row = score, r
            return best_row

        header_row = find_header_row()
        max_c = ws.max_column or 1
        headers = [norm(ws.cell(header_row, c).value) for c in range(1, max_c + 1)]
        header_map = {h: i + 1 for i, h in enumerate(headers) if h}

        def col(key):
            for alt in ALIASES.get(key, []):
                k = norm(alt)
                if k in header_map:
                    return header_map[k]
            return None

        def cellv(r, key):
            idx = col(key)
            if not idx:
                return None
            return ws.cell(r, idx).value

        # ======= counters =======
        added = updated = duplicated = skipped = 0

        # ======= loop data =======
        for r in range(header_row + 1, (ws.max_row or 1) + 1):
            emp_id = safe_str(cellv(r, "emp_id"))
            course_code = safe_str(cellv(r, "course_code"))
            start_date = safe_date(cellv(r, "start_date"))
            end_date = safe_date(cellv(r, "end_date"))

            prefix = safe_str(cellv(r, "prefix"))
            first_name = safe_str(cellv(r, "first_name"))
            last_name = safe_str(cellv(r, "last_name"))
            section = safe_str(cellv(r, "section"))
            position = safe_str(cellv(r, "position"))

            year = safe_int(cellv(r, "year"))
            month = safe_month(cellv(r, "month"))

            course_name = safe_str(cellv(r, "course_name"))
            course_type = safe_str(cellv(r, "course_type"))
            hours = safe_float(cellv(r, "hours"))
            evaluate_method = safe_str(cellv(r, "evaluate_method"))
            result = safe_str(cellv(r, "result"))
            score = safe_float(cellv(r, "score"))
            evaluator = safe_str(cellv(r, "evaluator"))
            expire_date = safe_date(cellv(r, "expire_date"))
            remark = safe_str(cellv(r, "remark"))

            # ---- ถ้าแถวว่างมาก ๆ ให้ข้ามแบบเงียบ ----
            if not emp_id and not course_code and not course_name and not start_date:
                continue

            # ---- SKIPPED: ข้อมูลขั้นต่ำไม่ครบ ----
            if not emp_id or not start_date or not end_date or not course_code:
                skipped += 1
                log_item(
                    "Skipped",
                    reason="ข้อมูลขั้นต่ำไม่ครบ (ต้องมี Emp ID + StartDate + EndDate + Course code)",
                    row_no=r,
                    emp_id=emp_id,
                    prefix=prefix,
                    first_name=first_name,
                    last_name=last_name,
                    section=section,
                    position=position,
                    course_code=course_code,
                    course_name=course_name,
                    course_type=course_type,
                    start_date=start_date,
                    end_date=end_date,
                )
                continue

            # ======================================================
            # KEY “ตัวเดิม” = emp_id + start_date + end_date + course_code
            # - ถ้า key นี้ไม่เคยมี → Added
            # - ถ้าเคยมี → พยายามเติมช่องว่าง (Updated) ไม่งั้น Duplicate
            # ======================================================
            existing = (
                TrainingRecord.query
                .filter(TrainingRecord.emp_id == emp_id)
                .filter(TrainingRecord.start_date == start_date)
                .filter(TrainingRecord.end_date == end_date)
                .filter(TrainingRecord.course_code == course_code)
                .first()
            )

            if existing is None:
                tr = TrainingRecord(
                    year=year,
                    month=month,
                    emp_id=emp_id,
                    prefix=prefix,
                    first_name=first_name,
                    last_name=last_name,
                    section=section,
                    position=position,
                    course_code=course_code,
                    course_name=course_name,
                    course_type=course_type,
                    start_date=start_date,
                    end_date=end_date,
                    hours=hours,
                    evaluate_method=evaluate_method,
                    result=result,
                    score=score,
                    evaluator=evaluator,
                    expire_date=expire_date,
                    remark=remark,
                )
                db.session.add(tr)
                added += 1
                log_item(
                    "Added",
                    row_no=r,
                    emp_id=emp_id,
                    prefix=prefix,
                    first_name=first_name,
                    last_name=last_name,
                    section=section,
                    position=position,
                    course_code=course_code,
                    course_name=course_name,
                    course_type=course_type,
                    start_date=start_date,
                    end_date=end_date,
                )
                continue

            # ---- มีตัวเดิมแล้ว: UPDATE เฉพาะช่องที่เดิมว่าง + ใหม่มีค่า ----
            updated_flag = False

            def fill_if_empty(field, new_value):
                nonlocal updated_flag
                old = getattr(existing, field)
                old_empty = (old is None) or (isinstance(old, str) and old.strip() == "")
                new_ok = (new_value is not None) and (not (isinstance(new_value, str) and new_value.strip() == ""))
                if old_empty and new_ok:
                    setattr(existing, field, new_value)
                    updated_flag = True
                    
            fill_if_empty("course_name", course_name)
            fill_if_empty("course_type", course_type)
            fill_if_empty("hours", hours)
            fill_if_empty("evaluate_method", evaluate_method)
            fill_if_empty("result", result)
            fill_if_empty("score", score)
            fill_if_empty("evaluator", evaluator)
            fill_if_empty("expire_date", expire_date)
            fill_if_empty("remark", remark)

            # ✅ แนะนำ: เติม prefix/name/section/position ได้ด้วย (เพราะตอนนี้ของคุณ first_name/section ว่างอยู่)
            fill_if_empty("prefix", prefix)
            fill_if_empty("first_name", first_name)
            fill_if_empty("last_name", last_name)
            fill_if_empty("section", section)
            fill_if_empty("position", position)
            fill_if_empty("year", year)
            fill_if_empty("month", month)

            if updated_flag:
                updated += 1
                log_item(
                    "Updated",
                    row_no=r,
                    emp_id=emp_id,
                    prefix=prefix,
                    first_name=first_name,
                    last_name=last_name,
                    section=section,
                    position=position,
                    course_code=course_code,
                    course_name=course_name,
                    course_type=course_type,
                    start_date=start_date,
                    end_date=end_date,
                )
            else:
                duplicated += 1
                log_item(
                    "Duplicate",
                    reason="พบรายการเดิมแล้ว และไม่มีข้อมูลใหม่เพื่อเติม",
                    row_no=r,
                    emp_id=emp_id,
                    prefix=prefix,
                    first_name=first_name,
                    last_name=last_name,
                    section=section,
                    position=position,
                    course_code=course_code,
                    course_name=course_name,
                    course_type=course_type,
                    start_date=start_date,
                    end_date=end_date,
                )

        # ---- บันทึกผลลง batch ----
        batch.added = added
        batch.updated = updated
        batch.duplicated = duplicated
        batch.skipped = skipped

        db.session.commit()
        audit("TRAINING_IMPORT", f"file={filename}, added={added}, updated={updated}, skipped={skipped}")
        
        flash(
            f"Import สำเร็จ: เพิ่ม {added} | อัปเดต {updated} | ซ้ำ {duplicated} | ข้าม {skipped}",
            "success"
        )
        return redirect(url_for("import_batch_detail", batch_id=batch.id))

    except Exception as e:
        db.session.rollback()
    
        flash(f"Import ไม่สำเร็จ: {e}", "error")
        return redirect(url_for("trainings_import"))


# =========================================================
# IMPORT HISTORY LIST
# =========================================================
@app.get("/training-imports")
def import_batches_list():
    batches = ImportBatch.query.order_by(ImportBatch.id.desc()).limit(50).all()
    return render_template("training-imports_list.html", batches=batches)


# =========================================================
# IMPORT HISTORY DETAIL
# =========================================================
@app.get("/training-imports/<int:batch_id>")
def import_batch_detail(batch_id):
    batch = ImportBatch.query.get_or_404(batch_id)

    added = ImportItem.query.filter_by(batch_id=batch_id, status="Added").order_by(ImportItem.id.asc()).all()
    updated = ImportItem.query.filter_by(batch_id=batch_id, status="Updated").order_by(ImportItem.id.asc()).all()
    duplicated = ImportItem.query.filter_by(batch_id=batch_id, status="Duplicate").order_by(ImportItem.id.asc()).all()
    skipped = ImportItem.query.filter_by(batch_id=batch_id, status="Skipped").order_by(ImportItem.id.asc()).all()

    return render_template(
        "training-imports_detail.html",
        batch=batch,
        added=added,
        updated=updated,
        duplicated=duplicated,
        skipped=skipped,
    )                    

@app.get("/admin/audit-logs")
def admin_audit_logs():
    u = get_current_user()
    if not u or u.role != "admin":
        abort(403)

    q = (request.args.get("q") or "").strip()

    query = AuditLog.query

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                AuditLog.user_email.ilike(like),
                AuditLog.action.ilike(like),
                AuditLog.detail.ilike(like),
                AuditLog.ip.ilike(like),
            )
        )

    rows = query.order_by(AuditLog.created_at.desc(), AuditLog.id.desc()).limit(500).all()

    return render_template("admin_audit_logs.html", rows=rows, q=q)

@app.get("/employees/export")
@login_required
def employees_export():
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "Active").strip()
    dept = (request.args.get("dept") or "").strip()
    section = (request.args.get("section") or "").strip()
    sort = (request.args.get("sort") or "no").strip()
    direction = (request.args.get("direction") or "asc").strip().lower()

    query = Employee.query

    if status in ["Active", "Resign"]:
        query = query.filter(Employee.status == status)

    if dept:
        query = query.filter(Employee.department == dept)
    if section:
        query = query.filter(Employee.section == section)

    if q:
        like = f"%{q}%"
        query = query.filter(
            or_(
                Employee.em_id.ilike(like),
                Employee.id_card.ilike(like),
                Employee.first_name_th.ilike(like),
                Employee.last_name_th.ilike(like),
                Employee.first_name_en.ilike(like),
                Employee.last_name_en.ilike(like),
                Employee.position.ilike(like),
                Employee.section.ilike(like),
                Employee.department.ilike(like),
                Employee.status.ilike(like),
            )
        )

    sort_map = {
        "no": Employee.no,
        "em_id": Employee.em_id,
        "department": Employee.department,
        "section": Employee.section,
    }
    sort_col = sort_map.get(sort, Employee.no)

    if direction == "desc":
        query = query.order_by(nullslast(sort_col.desc()))
    else:
        query = query.order_by(nullslast(sort_col.asc()))

    if sort != "em_id":
        query = query.order_by(
            nullslast(sort_col.desc() if direction == "desc" else sort_col.asc()),
            Employee.em_id.asc()
        )

    rows = query.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    ws.append([
        "No", "Em. ID", "ID Card", "ชื่อไทย", "Name-EN",
        "Position", "Section", "Department", "Start work",
        "Resign", "Status", "Degree", "Major"
    ])

    for e in rows:
        ws.append([
            e.no or "",
            e.em_id or "",
            e.id_card or "",
            e.th_full() if hasattr(e, "th_full") else f"{e.first_name_th or ''} {e.last_name_th or ''}".strip(),
            e.en_full() if hasattr(e, "en_full") else f"{e.first_name_en or ''} {e.last_name_en or ''}".strip(),
            e.position or "",
            e.section or "",
            e.department or "",
            str(e.start_work) if e.start_work else "",
            str(e.resign) if e.resign else "",
            e.status or "",
            e.degree or "",
            e.major or "",
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"employees_{status or 'All'}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.get("/dashboard")
def dashboard():
    # รับค่า year/month จาก query string เช่น /dashboard?year=2026&month=2
    today = date.today()
    year = request.args.get("year", type=int) or today.year
    month = request.args.get("month", type=int)  # อาจเป็น None

    # รวมทั้งหมด / active / resigned (ของทั้งระบบ)
    total = db.session.query(func.count(Employee.id)).scalar() or 0
    active = db.session.query(func.count(Employee.id)).filter(Employee.status == "W").scalar() or 0
    resigned_total = db.session.query(func.count(Employee.id)).filter(Employee.status == "RS").scalar() or 0

    # ----- เข้า/ออก "ปีนี้" -----
    joined_year = (
        db.session.query(func.count(Employee.id))
        .filter(Employee.start_work.isnot(None))
        .filter(func.extract("year", Employee.start_work) == year)
        .scalar()
        or 0
    )
    resigned_year = (
        db.session.query(func.count(Employee.id))
        .filter(Employee.resign.isnot(None))
        .filter(func.extract("year", Employee.resign) == year)
        .scalar()
        or 0
    )

    # ----- เข้า/ออก "เดือนนี้" (ถ้าเลือก month) -----
    joined_month = None
    resigned_month = None
    if month:
        joined_month = (
            db.session.query(func.count(Employee.id))
            .filter(Employee.start_work.isnot(None))
            .filter(func.extract("year", Employee.start_work) == year)
            .filter(func.extract("month", Employee.start_work) == month)
            .scalar()
            or 0
        )
        resigned_month = (
            db.session.query(func.count(Employee.id))
            .filter(Employee.resign.isnot(None))
            .filter(func.extract("year", Employee.resign) == year)
            .filter(func.extract("month", Employee.resign) == month)
            .scalar()
            or 0
        )

    # ----- สรุปเข้า/ออก รายเดือน (1-12) ของปีที่เลือก -----
    join_rows = (
        db.session.query(func.extract("month", Employee.start_work).label("m"), func.count(Employee.id))
        .filter(Employee.start_work.isnot(None))
        .filter(func.extract("year", Employee.start_work) == year)
        .group_by("m")
        .all()
    )
    resign_rows = (
        db.session.query(func.extract("month", Employee.resign).label("m"), func.count(Employee.id))
        .filter(Employee.resign.isnot(None))
        .filter(func.extract("year", Employee.resign) == year)
        .group_by("m")
        .all()
    )

    join_map = {int(m): c for m, c in join_rows if m is not None}
    resign_map = {int(m): c for m, c in resign_rows if m is not None}

    month_summary = []
    for m in range(1, 13):
        month_summary.append({
            "month": m,
            "joined": int(join_map.get(m, 0)),
            "resigned": int(resign_map.get(m, 0)),
        })

    start_date = date(year, month, 1) if month else date(year, 1, 1)
    end_date = date(year + (1 if month == 12 else 0), (1 if month == 12 else (month + 1)), 1) if month else date(year + 1, 1, 1)

# ---- เข้า: start_work ตามช่วง ----
    dept_join_rows = (
    db.session.query(Employee.section, func.count(Employee.id))
    .filter(Employee.start_work >= start_date, Employee.start_work < end_date)
    .group_by(Employee.section)
    .all()
    )

# ---- ออก: resign ตามช่วง ----
    dept_resign_rows = (
    db.session.query(Employee.section, func.count(Employee.id))
    .filter(Employee.resign >= start_date, Employee.resign < end_date)
    .group_by(Employee.section)
    .all()
    )

# map เป็น dict เพื่อ merge กันง่าย
    join_map = { (d or "ไม่ระบุ"): int(c) for d, c in dept_join_rows }
    resign_map = { (d or "ไม่ระบุ"): int(c) for d, c in dept_resign_rows }

# รวมเป็น list เดียว
    dept_inout = []
    all_depts = set(join_map.keys()) | set(resign_map.keys())
    for d in all_depts:
        dept_inout.append({
        "section": d,
        "joined": join_map.get(d, 0),
        "resigned": resign_map.get(d, 0),
        "net": join_map.get(d, 0) - resign_map.get(d, 0),
    })

# เรียง: ออกมากสุดก่อน หรือเข้าออกมากสุดก่อนก็ได้
    dept_inout.sort(key=lambda x: (x["joined"] + x["resigned"]), reverse=True)

# เอา Top 10
    dept_inout_top10 = dept_inout[:10]
    
    return render_template(
        "dashboard.html",
        total=total,
        active=active,
        resigned_total=resigned_total,
        resigned=resigned_total,
        year=year,
        month=month,
        joined_year=joined_year,
        resigned_year=resigned_year,
        joined_month=joined_month,
        resigned_month=resigned_month,
        month_summary=month_summary,
        dept_inout_top10=dept_inout_top10,
        start_date=start_date,
        end_date=end_date,
    )

@app.get("/trainings")
def trainings_list():
    query, q, year, month = build_training_query(request.args)

    # pagination
    page = int(request.args.get("page", 1) or 1)
    per_page = int(request.args.get("per_page", 50) or 50)
    if per_page not in (25, 50, 100, 200):
        per_page = 50
    if page < 1:
        page = 1

    total = query.count()
    total_pages = max(1, (total + per_page - 1) // per_page)

    rows = (query
        .order_by(TrainingRecord.start_date.desc().nullslast(), TrainingRecord.id.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    return render_template(
        "trainings_list.html",
        rows=rows,
        total=total,
        q=q, year=year, month=month,
        page=page,
        per_page=per_page,
        total_pages=total_pages,
    )

@app.get("/trainings/export")
@role_required("admin", "viewer")  # ✅ viewer export ได้
def trainings_export():
    query, q, year, month = build_training_query(request.args)

    # กัน export เยอะเกินจนล่ม (ปรับเลขได้)
    total = query.count()
    if total > 50000:
        flash(f"ผลลัพธ์ {total:,} รายการมากเกินไป กรุณากรองเพิ่มก่อน Export", "error")
        return redirect(url_for("trainings_list", q=q, year=year, month=month))

    # log สำหรับ audit
    audit("EXPORT_TRAININGS", f"q={q}&year={year}&month={month}&total={total}")

    # ใช้ write_only ลด memory
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("TrainingRecord")

    headers = [
        "Year","Month","Emp ID","Prefix","First Name","Last Name","Section","Position",
        "Course Code","Course Name","Type","StartDate","EndDate","Hours",
        "Eval Method","Result","Score","Evaluator","Expire Date","Remark"
    ]
    ws.append(headers)

    def val(x):
        if x is None:
            return ""
        # date/datetime -> string
        try:
            if hasattr(x, "strftime"):
                return x.strftime("%Y-%m-%d")
        except Exception:
            pass
        return str(x)

    # ดึงทีละ batch ลด RAM
    # (yield_per ใช้ได้กับหลาย DB; ถ้าเจอปัญหา ค่อยปรับ)
    for t in query.order_by(TrainingRecord.id.desc()).yield_per(1000):
        ws.append([
            val(getattr(t, "year", "")),
            val(getattr(t, "month", "")),
            val(getattr(t, "emp_id", getattr(t, "employee_code", ""))),
            val(getattr(t, "prefix", "")),
            val(getattr(t, "first_name", "")),
            val(getattr(t, "last_name", "")),
            val(getattr(t, "section", "")),
            val(getattr(t, "position", "")),
            val(getattr(t, "course_code", "")),
            val(getattr(t, "course_name", "")),
            val(getattr(t, "course_type", "")),
            val(getattr(t, "start_date", "")),
            val(getattr(t, "end_date", "")),
            val(getattr(t, "hours", "")),
            val(getattr(t, "evaluate_method", "")),
            val(getattr(t, "result", "")),
            val(getattr(t, "score", "")),
            val(getattr(t, "evaluator", "")),
            val(getattr(t, "expire_date", "")),
            val(getattr(t, "remark", "")),
        ])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"training_record_{stamp}.xlsx"

    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route("/trainings/<int:tr_id>/edit", methods=["GET", "POST"])
@role_required("admin")
def trainings_edit(tr_id):
    tr = TrainingRecord.query.get_or_404(tr_id)

    if request.method == "POST":
        tr.year = safe_int(request.form.get("year"))
        tr.month = safe_month(request.form.get("month"))

        tr.emp_id = safe_str(request.form.get("emp_id"))
        tr.prefix = safe_str(request.form.get("prefix"))
        tr.first_name = safe_str(request.form.get("first_name"))
        tr.last_name = safe_str(request.form.get("last_name"))

        tr.section = safe_str(request.form.get("section"))
        tr.position = safe_str(request.form.get("position"))

        tr.course_code = safe_str(request.form.get("course_code"))
        tr.course_name = safe_str(request.form.get("course_name"))
        tr.course_type = safe_str(request.form.get("course_type"))

        tr.start_date = safe_date(request.form.get("start_date"))
        tr.end_date = safe_date(request.form.get("end_date"))
        tr.hours = safe_float(request.form.get("hours"))

        tr.evaluate_method = safe_str(request.form.get("evaluate_method"))
        tr.result = safe_str(request.form.get("result"))
        tr.score = safe_float(request.form.get("score"))
        tr.evaluator = safe_str(request.form.get("evaluator"))
        tr.expire_date = safe_date(request.form.get("expire_date"))
        tr.remark = safe_str(request.form.get("remark"))

        db.session.commit()
        flash("แก้ไข Training Record เรียบร้อย", "success")
        return redirect(url_for("trainings_list"))

    return render_template("trainings_edit.html", tr=tr)

@app.route("/trainings/new", methods=["GET", "POST"])
def trainings_new():
    if request.method == "GET":
        return render_template("trainings_new.html")

    emp_id = safe_str(request.form.get("emp_id"))
    if not emp_id:
        flash("กรุณากรอก Emp ID", "error")
        return redirect(url_for("trainings_new"))

    tr = TrainingRecord(
        year=safe_int(request.form.get("year")),
        month=safe_int(request.form.get("month")),
        emp_id=emp_id,
        prefix=safe_str(request.form.get("prefix")),
        first_name=safe_str(request.form.get("first_name")),
        last_name=safe_str(request.form.get("last_name")),
        section=safe_str(request.form.get("section")),
        position=safe_str(request.form.get("position")),
        course_code=safe_str(request.form.get("course_code")),
        course_name=safe_str(request.form.get("course_name")),
        course_type=safe_str(request.form.get("course_type")),
        start_date=safe_date(request.form.get("start_date")),
        end_date=safe_date(request.form.get("end_date")),
        hours=safe_float(request.form.get("hours")),
        evaluate_method=safe_str(request.form.get("evaluate_method")),
        result=safe_str(request.form.get("result")),
        score=safe_float(request.form.get("score")),
        evaluator=safe_str(request.form.get("evaluator")),
        expire_date=safe_date(request.form.get("expire_date")),
        remark=safe_str(request.form.get("remark")),
    )

    db.session.add(tr)
    db.session.commit()
    flash("บันทึก Training Record แล้ว", "success")
    return redirect(url_for("trainings_list"))

@app.route("/trainings/bulk-delete", methods=["POST"])
@role_required("admin")
def trainings_bulk_delete():

    print("BULK DELETE IDS:", request.form.getlist("ids"))
    
    ids = request.form.getlist("ids")  # ✅ ต้อง getlist เท่านั้น

    if not ids:
        flash("ยังไม่ได้เลือกข้อมูลที่จะลบ", "error")
        return redirect(url_for("trainings_list"))

    try:
        # แปลงเป็น int กันค่ามั่ว
        ids_int = []
        for x in ids:
            try:
                ids_int.append(int(x))
            except:
                pass

        if not ids_int:
            flash("รายการที่เลือกไม่ถูกต้อง", "error")
            return redirect(url_for("trainings_list"))

        deleted = (
            TrainingRecord.query
            .filter(TrainingRecord.id.in_(ids_int))
            .delete(synchronize_session=False)
        )

        db.session.commit()
        flash(f"ลบสำเร็จ {deleted} รายการ", "success")
        return redirect(url_for("trainings_list"))

    except Exception as e:
        db.session.rollback()
        flash(f"ลบไม่สำเร็จ: {e}", "error")
        return redirect(url_for("trainings_list"))

@app.route("/trainings/<int:tr_id>/delete", methods=["POST"])
@role_required("admin")
def trainings_delete(tr_id):
    tr = TrainingRecord.query.get_or_404(tr_id)
    try:
        db.session.delete(tr)
        db.session.commit()
        flash("ลบรายการแล้ว", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"ลบไม่สำเร็จ: {e}", "error")
    return redirect(url_for("trainings_list"))

@app.get("/reports/training")
@login_required
def report_training():
    q = (request.args.get("q") or "").strip()

    emp = None
    rows = []

    if q:
        like = f"%{q}%"

        # หา employee จาก emp_id หรือชื่อ/นามสกุล (ไทย/อังกฤษ)
        emp = Employee.query.filter(
            or_(
                Employee.em_id.ilike(like),
                Employee.first_name_th.ilike(like),
                Employee.last_name_th.ilike(like),
                Employee.first_name_en.ilike(like),
                Employee.last_name_en.ilike(like),
            )
        ).order_by(Employee.em_id.asc()).first()

        # ถ้าไม่เจอใน Employee ให้ลองดึงจาก TrainingRecord โดยตรง
        if not emp:
            tr = TrainingRecord.query.filter(
                or_(
                    TrainingRecord.emp_id.ilike(like),
                    TrainingRecord.first_name.ilike(like),
                    TrainingRecord.last_name.ilike(like),
                )
            ).order_by(TrainingRecord.emp_id.asc()).first()

            if tr:
                class TempEmp:
                    def __init__(self, em_id, th_name, section, position):
                        self.em_id = em_id
                        self._th = th_name
                        self.section = section
                        self.position = position
                        self.department = None
                        self.start_work = None
                        self.resign = None
                        self.status = None

                    def th_full(self): return self._th
                    def en_full(self): return ""

                emp = TempEmp(
                    (tr.emp_id or "").strip(),
                    f"{tr.prefix or ''}{tr.first_name or ''} {tr.last_name or ''}".strip(),
                    tr.section,
                    tr.position,
                )

        # ✅ ดึง training record (match แบบ trim+upper)
        if emp:
            emp_id_norm = (emp.em_id or "").strip().upper()

            query = TrainingRecord.query.filter(
                func.upper(func.trim(TrainingRecord.emp_id)) == emp_id_norm
            ).order_by(
                nullslast(TrainingRecord.start_date.desc()),
                TrainingRecord.id.desc()
            )

            rows = query.all()

    return render_template(
        "report_training.html",
        q=q,
        emp=emp,
        rows=rows,
    )


@app.get("/reports/training/print")
@login_required
def report_training_print():
    emp_id = (request.args.get("emp_id") or "").strip()
    if not emp_id:
        flash("กรุณาระบุ Emp ID", "error")
        return redirect(url_for("report_training"))

    emp_id_norm = emp_id.strip().upper()

    # employee อาจมีหรือไม่มีก็ได้
    # ✅ แนะนำ normalize ตอนหา employee ด้วย (กัน em_id มีช่องว่าง)
    emp = Employee.query.filter(
        func.upper(func.trim(Employee.em_id)) == emp_id_norm
    ).first()

    # fallback: ถ้าไม่มี employee ให้ใช้ข้อมูลจาก training_records
    if not emp:
        tr = TrainingRecord.query.filter(
            func.upper(func.trim(TrainingRecord.emp_id)) == emp_id_norm
        ).order_by(TrainingRecord.id.desc()).first()

        if not tr:
            flash("ไม่พบข้อมูล Training Record ของพนักงานนี้", "error")
            return redirect(url_for("report_training", q=emp_id))

        class TempEmp:
            def __init__(self, em_id, th_name, section, position):
                self.em_id = em_id
                self._th = th_name
                self.section = section
                self.position = position
                self.department = None
                self.start_work = None
                self.resign = None
                self.status = None

            def th_full(self): return self._th
            def en_full(self): return ""

        emp = TempEmp(
            (tr.emp_id or "").strip(),
            f"{tr.prefix or ''}{tr.first_name or ''} {tr.last_name or ''}".strip(),
            tr.section,
            tr.position,
        )

    # ✅ ดึงข้อมูล training ทั้งหมดแบบ trim+upper (สำคัญมาก)
    rows = TrainingRecord.query.filter(
        func.upper(func.trim(TrainingRecord.emp_id)) == emp_id_norm
    ).order_by(
        nullslast(TrainingRecord.start_date.asc()),
        TrainingRecord.id.asc()
    ).all()

    # ✅ แบ่งหน้า: 10 รายการ/หน้า
    per_page = 10
    total_pages = max(1, ceil(len(rows) / per_page))
    pages = [rows[i:i + per_page] for i in range(0, len(rows), per_page)]

    return render_template(
        "report_training_print.html",
        emp=emp,
        rows=rows,               # เผื่อใช้
        pages=pages,             # ✅ ใช้ทำหลายหน้า
        per_page=per_page,
        total_pages=total_pages,
        print_date=datetime.utcnow(),
    )

@app.get("/reports/training/export")
@login_required
def report_training_export():
    emp_id = (request.args.get("emp_id") or "").strip()
    if not emp_id:
        flash("กรุณาระบุ Emp ID", "error")
        return redirect(url_for("report_training"))

    # normalize ให้ match เหมือนที่คุณแก้ใน report_training
    emp_id_norm = emp_id.strip().upper()

    rows = TrainingRecord.query.filter(
        func.upper(func.trim(TrainingRecord.emp_id)) == emp_id_norm
    ).order_by(
        nullslast(TrainingRecord.start_date.asc()),
        TrainingRecord.id.asc()
    ).all()

    if not rows:
        flash("ไม่พบ Training Record สำหรับ Emp ID นี้", "error")
        return redirect(url_for("report_training", q=emp_id))

    # employee (มี/ไม่มีได้)
    emp = Employee.query.filter(func.upper(func.trim(Employee.em_id)) == emp_id_norm).first()

    wb = Workbook()
    ws = wb.active
    ws.title = "Training Report"

    # ===== Header บนสุด =====
    ws["A1"] = "Training Report"
    ws["A2"] = f"Emp ID: {emp_id_norm}"
    if emp:
        ws["A3"] = f"Name-TH: {emp.th_full()}"
        ws["A4"] = f"Section: {emp.section or '-'}   Position: {emp.position or '-'}"
        if getattr(emp, "department", None):
            ws["A5"] = f"Department: {emp.department}"
    ws["A6"] = " "

    # ===== Table header =====
    headers = [
        "#", "Year", "Month",
        "Course Code", "Course Name", "Type",
        "Start Date", "End Date", "Hours",
        "Eval Method", "Result", "Score",
        "Evaluator", "Expire Date", "Remark"
    ]
    start_row = 7
    ws.append([])  # ให้ row 7 อยู่
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=col_idx, value=h)

    # ===== Table rows =====
    r = start_row + 1
    for i, t in enumerate(rows, start=1):
        ws.cell(r, 1, i)
        ws.cell(r, 2, t.year)
        ws.cell(r, 3, t.month)
        ws.cell(r, 4, t.course_code)
        ws.cell(r, 5, t.course_name)
        ws.cell(r, 6, t.course_type)
        ws.cell(r, 7, t.start_date.isoformat() if t.start_date else "")
        ws.cell(r, 8, t.end_date.isoformat() if t.end_date else "")
        ws.cell(r, 9, t.hours)
        ws.cell(r, 10, t.evaluate_method)
        ws.cell(r, 11, t.result)
        ws.cell(r, 12, t.score)
        ws.cell(r, 13, t.evaluator)
        ws.cell(r, 14, t.expire_date.isoformat() if t.expire_date else "")
        ws.cell(r, 15, t.remark)
        r += 1

    # ===== Auto width แบบง่าย =====
    for col in range(1, 16):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["O"].width = 30

    # ===== ส่งไฟล์ =====
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"TrainingReport_{emp_id_norm}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.get("/courses")
@login_required
def courses_list():
    q = (request.args.get("q") or "").strip()
    ctype = (request.args.get("type") or "All").strip()

    query = TrainingCourse.query

    if ctype in ["OJT", "INH", "EXT"]:
        query = query.filter(TrainingCourse.course_type == ctype)

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            TrainingCourse.course_code.ilike(like),
            TrainingCourse.course_name.ilike(like),
            TrainingCourse.vendor.ilike(like),
            TrainingCourse.owner.ilike(like),
        ))

    rows = query.order_by(TrainingCourse.created_at.desc()).all()
    return render_template("courses_list.html", rows=rows, q=q, ctype=ctype)


@app.route("/courses/new", methods=["GET", "POST"])
@login_required
@role_required("admin")
def course_new():
    if request.method == "GET":
        return render_template("course_form.html", mode="new", course=None)

    course_type = (request.form.get("course_type") or "").strip().upper()
    course_name = (request.form.get("course_name") or "").strip()

    if course_type not in ["OJT", "INH", "EXT"]:
        flash("กรุณาเลือกประเภท OJT / INH / EXT", "error")
        return redirect(url_for("course_new"))

    if not course_name:
        flash("กรุณากรอกชื่อหลักสูตร", "error")
        return redirect(url_for("course_new"))

    now = datetime.utcnow()
    owner = (request.form.get("owner") or "").strip()
    code = gen_course_code(course_type, owner, now)
    training_hours_raw = (request.form.get("training_hours") or "").strip()
    try:
        training_hours = float(training_hours_raw) if training_hours_raw else None
    except Exception:
        training_hours = None

    c = TrainingCourse(
        course_type=course_type,
        course_code=code,
        course_name=course_name,
        description=(request.form.get("description") or "").strip() or None,
        owner=(request.form.get("owner") or "").strip() or None,
        vendor=(request.form.get("vendor") or "").strip() or None,
        location=(request.form.get("location") or "").strip() or None,
        course_year=now.year,
        course_month=now.month,
        status=(request.form.get("status") or "Draft").strip() or "Draft",
    )

    db.session.add(c)
    db.session.commit()

    audit("COURSE_ADD", f"course_code={c.course_code}")
    flash(f"สร้างหลักสูตรสำเร็จ: {code}", "success")

    return redirect(url_for("course_edit", course_id=c.id))

@app.route("/courses/<int:course_id>/edit", methods=["GET", "POST"])
@login_required
@role_required("admin")
def course_edit(course_id):
    c = db.session.get(TrainingCourse, course_id)
    if not c:
        flash("ไม่พบหลักสูตร", "error")
        return redirect(url_for("courses_list"))

    if request.method == "POST":
        c.course_name = (request.form.get("course_name") or "").strip()
        c.description = (request.form.get("description") or "").strip() or None
        c.owner = (request.form.get("owner") or "").strip() or None
        c.vendor = (request.form.get("vendor") or "").strip() or None
        c.location = (request.form.get("location") or "").strip() or None
        c.status = (request.form.get("status") or "Draft").strip() or "Draft"

        training_hours_raw = (request.form.get("training_hours") or "").strip()
        try:
            c.training_hours = float(training_hours_raw) if training_hours_raw else None
        except Exception:
            c.training_hours = None
        
        db.session.commit()

        audit("COURSE_EDIT", f"course_code={c.course_code}")
        flash("บันทึกข้อมูลหลักสูตรแล้ว", "success")
        return redirect(url_for("course_edit", course_id=course_id))

    return render_template("course_form.html", mode="edit", course=c)


@app.post("/courses/<int:course_id>/cost/add")
@login_required
@role_required("admin")
def course_cost_add(course_id):
    c = db.session.get(TrainingCourse, course_id)
    if not c:
        flash("ไม่พบหลักสูตร", "error")
        return redirect(url_for("courses_list"))

    cost_type = (request.form.get("cost_type") or "").strip()
    before = request.form.get("amount_before_vat") or ""
    vat_rate = request.form.get("vat_rate") or "7"

    try:
        before_f = float(before) if before != "" else 0.0
        vat_rate_f = float(vat_rate) if vat_rate != "" else 7.0
    except Exception:
        flash("จำนวนเงินไม่ถูกต้อง", "error")
        return redirect(url_for("course_edit", course_id=course_id))

    vat_amt = round(before_f * (vat_rate_f / 100.0), 2)
    total_amt = round(before_f + vat_amt, 2)

    item = CourseCostItem(
        course_id=course_id,
        cost_type=cost_type or "อื่นๆ",
        amount_before_vat=before_f,
        vat_rate=vat_rate_f,
        amount_vat=vat_amt,
        amount_total=total_amt,
        remark=(request.form.get("remark") or "").strip() or None,
    )

    db.session.add(item)
    db.session.commit()

    audit("COURSE_COST_ADD", f"course_code={c.course_code}, cost_type={item.cost_type}, total={item.amount_total}")
    flash("เพิ่มค่าใช้จ่ายแล้ว", "success")

    return redirect(url_for("course_edit", course_id=course_id))

@app.post("/courses/<int:course_id>/file/add")
@login_required
@role_required("admin")
def course_file_add(course_id):
    c = db.session.get(TrainingCourse, course_id)
    if not c:
        flash("ไม่พบหลักสูตร", "error")
        return redirect(url_for("courses_list"))

    f = request.files.get("file")
    if not f or not f.filename:
        flash("กรุณาเลือกไฟล์", "error")
        return redirect(url_for("course_edit", course_id=course_id))

    if not allowed_file(f.filename):
        flash("อนุญาตเฉพาะ pdf/png/jpg/jpeg/xlsx", "error")
        return redirect(url_for("course_edit", course_id=course_id))

    original = f.filename
    safe = secure_filename(original)
    stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    stored = f"{c.course_code}_{stamp}_{safe}"

    save_path = os.path.join(UPLOAD_COURSE_DIR, stored)
    f.save(save_path)

    cf = CourseFile(
        course_id=course_id,
        file_type=(request.form.get("file_type") or "").strip() or None,
        original_name=original,
        stored_name=stored,
        note=(request.form.get("note") or "").strip() or None,
    )

    db.session.add(cf)
    db.session.commit()

    audit("COURSE_FILE_ADD", f"course_code={c.course_code}, file={cf.original_name}")
    flash("แนบไฟล์แล้ว", "success")

    return redirect(url_for("course_edit", course_id=course_id))

@app.get("/events")
@login_required
def events_list():
    rows = TrainingEvent.query.order_by(TrainingEvent.created_at.desc()).all()
    return render_template("events_list.html", rows=rows)

@app.route("/events/new", methods=["GET", "POST"])
@login_required
@role_required("admin")
def events_new():
    courses = TrainingCourse.query.order_by(TrainingCourse.course_name.asc()).all()

    if request.method == "GET":
        return render_template("events_new.html", courses=courses)

    course_id = request.form.get("course_id")
    title = (request.form.get("title") or "").strip()
    location = (request.form.get("location") or "").strip()
    trainer = (request.form.get("trainer") or "").strip()
    start_date_raw = (request.form.get("start_date") or "").strip()
    end_date_raw = (request.form.get("end_date") or "").strip()

    if not course_id or not start_date_raw:
        flash("กรุณากรอกข้อมูลให้ครบ", "error")
        return redirect(url_for("events_new"))

    course = TrainingCourse.query.get_or_404(int(course_id))

    # ดึงค่าจาก Course ถ้ายังไม่ได้กรอกเอง
    if not trainer:
        trainer = (course.vendor or "").strip()

    if not location:
        location = (course.location or "").strip()

    event_type = (course.course_type or "").strip().upper()
    if event_type not in ["OJT", "INH", "EXT"]:
        flash("ประเภทหลักสูตรของ Course ไม่ถูกต้อง", "error")
        return redirect(url_for("events_new"))

    start_date = datetime.strptime(start_date_raw, "%Y-%m-%d").date()
    end_date = datetime.strptime(end_date_raw, "%Y-%m-%d").date() if end_date_raw else None

    code = gen_event_code(event_type, start_date)

    if not title:
        title = course.course_name

    e = TrainingEvent(
        course_id=course.id,
        event_type=event_type,
        event_code=code,
        title=title,
        location=location or None,
        trainer=trainer or None,
        start_date=start_date,
        end_date=end_date,
        status="PLANNED",
        description=course.description,
    )

    db.session.add(e)
    db.session.commit()

    audit("EVENT_CREATE", f"event_code={e.event_code}, course_code={course.course_code}")
    flash("สร้าง Training Event สำเร็จ", "success")

    return redirect(url_for("event_detail", event_id=e.id))

@app.get("/events/<int:event_id>")
@login_required
def event_detail(event_id):
    event = TrainingEvent.query.get_or_404(event_id)

    participant_rows = TrainingEventParticipant.query.filter_by(
        event_id=event.id
    ).order_by(TrainingEventParticipant.id.asc()).all()

    participants = []
    for p in participant_rows:
        emp = Employee.query.filter_by(em_id=p.emp_id).first()
        participants.append({
            "row": p,
            "emp": emp
        })

    total_before_vat = sum((x.amount_before_vat or 0) for x in event.cost_items)
    total_vat = sum((x.amount_vat or 0) for x in event.cost_items)
    total_amount = sum((x.amount_total or 0) for x in event.cost_items)

    return render_template(
        "event_detail.html",
        event=event,
        participants=participants,
        total_before_vat=total_before_vat,
        total_vat=total_vat,
        total_amount=total_amount,
        employees=employees,
    )


@app.post("/events/participant/<int:participant_id>/update")
@login_required
@role_required("admin")
def event_participant_update(participant_id):
    p = TrainingEventParticipant.query.get_or_404(participant_id)

    result = (request.form.get("result") or "").strip().upper()
    score_raw = (request.form.get("score") or "").strip()
    hours_raw = (request.form.get("training_hours") or "").strip()
    remark = (request.form.get("remark") or "").strip()

    p.result = result or None

    try:
        p.score = float(score_raw) if score_raw else None
    except Exception:
        p.score = None

    try:
        p.training_hours = float(hours_raw) if hours_raw else None
    except Exception:
        p.training_hours = None

    p.remark = remark or None

    db.session.commit()

    audit("EVENT_PARTICIPANT_UPDATE", f"event_id={p.event_id}, emp_id={p.emp_id}, result={p.result}")
    flash("บันทึกผลอบรมแล้ว", "success")

    return redirect(url_for("event_detail", event_id=p.event_id))


@app.post("/events/<int:event_id>/participants/add")
@login_required
@role_required("admin")
def event_participant_add(event_id):
    event = TrainingEvent.query.get_or_404(event_id)

    emp_id = (request.form.get("emp_id") or "").strip()

    if not emp_id:
        flash("กรุณาระบุ Emp ID", "error")
        return redirect(url_for("event_detail", event_id=event.id))

    emp = Employee.query.filter_by(em_id=emp_id).first()
    if not emp:
        flash("ไม่พบ Emp ID นี้ในระบบพนักงาน", "error")
        return redirect(url_for("event_detail", event_id=event.id))

    exists = TrainingEventParticipant.query.filter_by(
        event_id=event.id,
        emp_id=emp_id
    ).first()

    if exists:
        flash("พนักงานคนนี้อยู่ใน Event แล้ว", "error")
        return redirect(url_for("event_detail", event_id=event.id))

    row = TrainingEventParticipant(
        event_id=event.id,
        emp_id=emp_id
    )

    db.session.add(row)
    db.session.commit()

    audit("EVENT_PARTICIPANT_ADD", f"event_code={event.event_code}, emp_id={emp_id}")
    flash("เพิ่มผู้เข้าอบรมสำเร็จ", "success")

    return redirect(url_for("event_detail", event_id=event.id))


@app.post("/events/participants/<int:participant_id>/delete")
@login_required
@role_required("admin")
def event_participant_delete(participant_id):
    row = TrainingEventParticipant.query.get_or_404(participant_id)

    event_id = row.event_id
    emp_id = row.emp_id

    db.session.delete(row)
    db.session.commit()

    audit("EVENT_PARTICIPANT_DELETE", f"event_id={event_id}, emp_id={emp_id}")
    flash("ลบผู้เข้าอบรมออกจาก Event แล้ว", "success")

    return redirect(url_for("event_detail", event_id=event_id))


@app.post("/events/<int:event_id>/files/add")
@login_required
@role_required("admin")
def event_file_add(event_id):
    event = TrainingEvent.query.get_or_404(event_id)

    f = request.files.get("file")
    file_type = (request.form.get("file_type") or "").strip()
    note = (request.form.get("note") or "").strip()

    if not f or not f.filename:
        flash("กรุณาเลือกไฟล์", "error")
        return redirect(url_for("event_detail", event_id=event.id))

    if not allowed_file(f.filename):
        flash("อนุญาตเฉพาะไฟล์ pdf / png / jpg / jpeg / xlsx", "error")
        return redirect(url_for("event_detail", event_id=event.id))

    original = f.filename
    safe = secure_filename(original)
    stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    stored = f"{event.event_code}_{stamp}_{safe}"

    save_path = os.path.join(UPLOAD_EVENT_DIR, stored)
    f.save(save_path)

    row = EventFile(
        event_id=event.id,
        file_type=file_type or "other",
        original_name=original,
        stored_name=stored,
        note=note or None,
    )

    db.session.add(row)
    db.session.commit()

    audit("EVENT_FILE_ADD", f"event_code={event.event_code}, file={original}")
    flash("อัปโหลดไฟล์สำเร็จ", "success")

    return redirect(url_for("event_detail", event_id=event.id))


@app.post("/events/files/<int:file_id>/edit")
@login_required
@role_required("admin")
def event_file_edit(file_id):
    row = EventFile.query.get_or_404(file_id)

    row.file_type = (request.form.get("file_type") or "").strip() or "other"
    row.note = (request.form.get("note") or "").strip() or None

    db.session.commit()

    audit("EVENT_FILE_EDIT", f"file_id={row.id}, event_id={row.event_id}")
    flash("อัปเดตข้อมูลไฟล์แล้ว", "success")

    return redirect(url_for("event_detail", event_id=row.event_id))


@app.post("/events/files/<int:file_id>/delete")
@login_required
@role_required("admin")
def event_file_delete(file_id):
    row = EventFile.query.get_or_404(file_id)

    event_id = row.event_id
    original_name = row.original_name
    file_path = os.path.join(UPLOAD_EVENT_DIR, row.stored_name)

    try:
        if os.path.exists(file_path):
            os.remove(file_path)
    except Exception:
        pass

    db.session.delete(row)
    db.session.commit()

    audit("EVENT_FILE_DELETE", f"event_id={event_id}, file={original_name}")
    flash("ลบไฟล์แล้ว", "success")

    return redirect(url_for("event_detail", event_id=event_id))


@app.post("/events/<int:event_id>/cost/add")
@login_required
@role_required("admin")
def event_cost_add(event_id):
    event = TrainingEvent.query.get_or_404(event_id)

    cost_type = (request.form.get("cost_type") or "").strip()
    remark = (request.form.get("remark") or "").strip()

    try:
        amount_before_vat = float(request.form.get("amount_before_vat") or 0)
    except Exception:
        amount_before_vat = 0.0

    try:
        vat_rate = float(request.form.get("vat_rate") or 7)
    except Exception:
        vat_rate = 7.0

    if not cost_type:
        flash("กรุณาระบุประเภทค่าใช้จ่าย", "error")
        return redirect(url_for("event_detail", event_id=event.id))

    amount_vat = round(amount_before_vat * vat_rate / 100, 2)
    amount_total = round(amount_before_vat + amount_vat, 2)

    row = EventCostItem(
        event_id=event.id,
        cost_type=cost_type,
        amount_before_vat=amount_before_vat,
        vat_rate=vat_rate,
        amount_vat=amount_vat,
        amount_total=amount_total,
        remark=remark or None,
    )

    db.session.add(row)
    db.session.commit()

    audit("EVENT_COST_ADD", f"event_code={event.event_code}, cost_type={cost_type}, total={amount_total}")
    flash("เพิ่มค่าใช้จ่ายสำเร็จ", "success")

    return redirect(url_for("event_detail", event_id=event.id))


@app.post("/events/cost/<int:cost_id>/edit")
@login_required
@role_required("admin")
def event_cost_edit(cost_id):
    row = EventCostItem.query.get_or_404(cost_id)

    cost_type = (request.form.get("cost_type") or "").strip()
    remark = (request.form.get("remark") or "").strip()

    try:
        amount_before_vat = float(request.form.get("amount_before_vat") or 0)
    except Exception:
        amount_before_vat = 0.0

    try:
        vat_rate = float(request.form.get("vat_rate") or 7)
    except Exception:
        vat_rate = 7.0

    if not cost_type:
        flash("กรุณาระบุประเภทค่าใช้จ่าย", "error")
        return redirect(url_for("event_detail", event_id=row.event_id))

    amount_vat = round(amount_before_vat * vat_rate / 100, 2)
    amount_total = round(amount_before_vat + amount_vat, 2)

    row.cost_type = cost_type
    row.amount_before_vat = amount_before_vat
    row.vat_rate = vat_rate
    row.amount_vat = amount_vat
    row.amount_total = amount_total
    row.remark = remark or None

    db.session.commit()

    audit("EVENT_COST_EDIT", f"event_id={row.event_id}, cost_id={row.id}, total={amount_total}")
    flash("แก้ไขค่าใช้จ่ายแล้ว", "success")

    return redirect(url_for("event_detail", event_id=row.event_id))


@app.post("/events/cost/<int:cost_id>/delete")
@login_required
@role_required("admin")
def event_cost_delete(cost_id):
    row = EventCostItem.query.get_or_404(cost_id)

    event_id = row.event_id
    total = row.amount_total
    cost_type = row.cost_type

    db.session.delete(row)
    db.session.commit()

    audit("EVENT_COST_DELETE", f"event_id={event_id}, cost_type={cost_type}, total={total}")
    flash("ลบค่าใช้จ่ายแล้ว", "success")

    return redirect(url_for("event_detail", event_id=event_id))


@app.post("/events/<int:event_id>/generate-records")
@login_required
@role_required("admin")
def event_generate_training_records(event_id):
    event = TrainingEvent.query.get_or_404(event_id)

    participants = TrainingEventParticipant.query.filter_by(
        event_id=event.id
    ).order_by(TrainingEventParticipant.id.asc()).all()

    if not participants:
        flash("ยังไม่มีผู้เข้าอบรมใน Event นี้", "error")
        return redirect(url_for("event_detail", event_id=event.id))

    added = 0
    skipped = 0

    for p in participants:
        emp = Employee.query.filter_by(em_id=p.emp_id).first()

        if not emp:
            skipped += 1
            continue

        exists = TrainingRecord.query.filter_by(
            emp_id=p.emp_id,
            event_id=event.id
        ).first()

        if exists:
            skipped += 1
            continue

        rec = TrainingRecord(
            year=event.start_date.year if event.start_date else None,
            month=event.start_date.month if event.start_date else None,

            emp_id=p.emp_id,

            prefix=emp.title_th,
            first_name=emp.first_name_th,
            last_name=emp.last_name_th,

            section=emp.section,
            position=emp.position,

            course_code=event.course.course_code,
            course_name=event.course.course_name,
            course_type=event.course.course_type,

            start_date=event.start_date,
            end_date=event.end_date,

            hours=p.training_hours,

            evaluate_method="Event Result",
            result=p.result,
            score=p.score,
            evaluator=event.trainer,

            remark=p.remark,
            event_id=event.id
        )

        db.session.add(rec)
        added += 1

    db.session.commit()

    audit("GENERATE_TRAINING_RECORDS", f"event_code={event.event_code}, added={added}, skipped={skipped}")
    flash(f"สร้าง TrainingRecord สำเร็จ {added} รายการ / ข้าม {skipped} รายการ", "success")

    return redirect(url_for("event_detail", event_id=event.id))

@app.post("/courses/file/<int:file_id>/delete")
@login_required
@role_required("admin")
def course_file_delete(file_id):
    f = db.session.get(CourseFile, file_id)

    if not f:
        flash("ไม่พบไฟล์", "error")
        return redirect(url_for("courses_list"))

    course_id = f.course_id

    try:
        path = os.path.join(UPLOAD_COURSE_DIR, f.stored_name)
        if os.path.exists(path):
            os.remove(path)
    except Exception:
        pass

    db.session.delete(f)
    db.session.commit()

    audit("COURSE_FILE_DELETE", f"file_id={file_id}, course_id={course_id}")
    flash("ลบไฟล์แล้ว", "success")

    return redirect(url_for("course_edit", course_id=course_id))

@app.post("/courses/<int:course_id>/delete")
@login_required
@role_required("admin")
def course_delete(course_id):
    c = db.session.get(TrainingCourse, course_id)

    if not c:
        flash("ไม่พบหลักสูตร", "error")
        return redirect(url_for("courses_list"))

    # ถ้ามี Event อยู่ ห้ามลบ
    event_count = TrainingEvent.query.filter_by(course_id=c.id).count()
    if event_count > 0:
        flash("ไม่สามารถลบหลักสูตรได้ เพราะมี Training Event ใช้งานอยู่", "error")
        return redirect(url_for("course_edit", course_id=c.id))

    course_code = c.course_code

    # 1) ลบไฟล์จริงก่อน
    files = CourseFile.query.filter_by(course_id=c.id).all()
    for f in files:
        try:
            path = os.path.join(UPLOAD_COURSE_DIR, f.stored_name)
            if os.path.exists(path):
                os.remove(path)
        except Exception:
            pass

    # 2) ลบ rows ลูกทั้งหมดก่อน
    CourseFile.query.filter_by(course_id=c.id).delete(synchronize_session=False)
    CourseCostItem.query.filter_by(course_id=c.id).delete(synchronize_session=False)

    # ถ้ามี relation อื่นในอนาคตค่อยเพิ่มตรงนี้

    # 3) expunge/refresh session แล้วค่อยลบ course
    db.session.flush()

    c = db.session.get(TrainingCourse, course_id)
    if c:
        db.session.delete(c)

    db.session.commit()

    audit("COURSE_DELETE", f"course_code={course_code}")
    flash("ลบหลักสูตรแล้ว", "success")
    return redirect(url_for("courses_list"))
    
@app.post("/events/<int:event_id>/delete")
@login_required
@role_required("admin")
def event_delete(event_id):

    e = TrainingEvent.query.get_or_404(event_id)

    code = e.event_code

    # ลบ participant ก่อน
    TrainingEventParticipant.query.filter_by(event_id=e.id).delete()

    # ลบ files
    files = EventFile.query.filter_by(event_id=e.id).all()
    for f in files:
        try:
            path = os.path.join(UPLOAD_EVENT_DIR, f.stored_name)
            if os.path.exists(path):
                os.remove(path)
        except:
            pass
        db.session.delete(f)

    # ลบ cost
    EventCostItem.query.filter_by(event_id=e.id).delete()

    db.session.delete(e)
    db.session.commit()

    audit("EVENT_DELETE", f"event_code={code}")

    flash("ลบ Training Event แล้ว", "success")

    return redirect(url_for("events_list"))

@app.get("/events/calendar")
@login_required
def events_calendar():
    return render_template("events_calendar.html")
    
@app.get("/api/events/calendar")
@login_required
def events_calendar_api():
    rows = TrainingEvent.query.order_by(TrainingEvent.start_date.asc()).all()

    data = []

    for e in rows:
        if e.event_type == "OJT":
            color = "#f59e0b"
        elif e.event_type == "INH":
            color = "#2563eb"
        elif e.event_type == "EXT":
            color = "#16a34a"
        else:
            color = "#6b7280"

        data.append({
            "id": e.id,
            "title": f"{e.event_code} • {e.title}",
            "start": e.start_date.isoformat() if e.start_date else None,
            "end": e.end_date.isoformat() if e.end_date else None,
            "url": url_for("event_detail", event_id=e.id),
            "backgroundColor": color,
            "borderColor": color,
        })

    return data

@app.post("/events/<int:event_id>/participants/pass-all")
@login_required
@role_required("admin")
def event_participants_pass_all(event_id):
    event = TrainingEvent.query.get_or_404(event_id)

    rows = TrainingEventParticipant.query.filter_by(event_id=event.id).all()
    for r in rows:
        r.result = "PASS"

    db.session.commit()
    flash("ตั้งค่า PASS ให้ทุกคนแล้ว", "success")
    return redirect(url_for("event_detail", event_id=event.id))

@app.post("/events/<int:event_id>/participants/fill-hours")
@login_required
@role_required("admin")
def event_participants_fill_hours(event_id):
    event = TrainingEvent.query.get_or_404(event_id)

    hours = event.course.training_hours
    if hours is None:
        flash("Course นี้ยังไม่ได้กำหนด Training Hours", "error")
        return redirect(url_for("event_detail", event_id=event.id))

    rows = TrainingEventParticipant.query.filter_by(event_id=event.id).all()
    for r in rows:
        r.training_hours = hours

    db.session.commit()
    flash("เติมชั่วโมงจาก Course ให้ทุกคนแล้ว", "success")
    return redirect(url_for("event_detail", event_id=event.id))

@app.post("/events/<int:event_id>/participants/save-all")
@login_required
@role_required("admin")
def event_participants_save_all(event_id):
    event = TrainingEvent.query.get_or_404(event_id)

    rows = TrainingEventParticipant.query.filter_by(event_id=event.id).all()

    for r in rows:
        result = (request.form.get(f"result_{r.id}") or "").strip().upper()
        score_raw = (request.form.get(f"score_{r.id}") or "").strip()
        hours_raw = (request.form.get(f"hours_{r.id}") or "").strip()
        remark = (request.form.get(f"remark_{r.id}") or "").strip()

        r.result = result or None

        try:
            r.score = float(score_raw) if score_raw else None
        except Exception:
            r.score = None

        try:
            r.training_hours = float(hours_raw) if hours_raw else None
        except Exception:
            r.training_hours = None

        r.remark = remark or None

    db.session.commit()
    flash("บันทึกข้อมูลผู้เข้าอบรมทั้งหมดแล้ว", "success")
    return redirect(url_for("event_detail", event_id=event.id))
    
# -------------------------------------------------
# Run (Local Only)
# -------------------------------------------------
    with app.app_context():
        db.create_all()
        seed_users_if_missing()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
